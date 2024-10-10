# import pandas as pd
# import time  # For adding delays
# from googletrans import Translator
# from requests.exceptions import Timeout, RequestException

# # Initialize the translator
# translator = Translator()

# # Load the Excel file
# file_path = 'text_input.xlsx'  # Change this to your file path
# df = pd.read_excel(file_path, engine='openpyxl')

# # Function to check if text contains Malayalam characters
# def contains_malayalam(text):
#     if isinstance(text, str):
#         return any('\u0D00' <= char <= '\u0D7F' for char in text)
#     return False

# # Function to translate text if it contains Malayalam with retry logic
# def translate_text(text, retries=3, delay=5):
#     if contains_malayalam(text):
#         for attempt in range(retries):
#             try:
#                 translated = translator.translate(text, src='ml', dest='en')
#                 return translated.text
#             except (Timeout, RequestException) as e:
#                 print(f"Error translating: {text}. Error: {e}. Attempt {attempt + 1}/{retries}")
#                 time.sleep(delay)  # Wait before retrying
#             except Exception as e:
#                 print(f"Unexpected error for text: {text}. Error: {e}")
#                 break
#         return text  # Return original if translation fails after retries
#     return text  # Return original if not in Malayalam

# # Process each row individually
# for idx, row in df.iterrows():
#     # Iterate over each cell in the row
#     for col in df.columns:
#         cell_value = row[col]
#         df.at[idx, col] = translate_text(cell_value)  # Translate the text in the cell if necessary

#     # Display a success message for each row processed
#     print(f"Row {idx + 1} processed successfully.")

# # Save the translated DataFrame to a new Excel file
# output_file = 'translated_file.xlsx'
# df.to_excel(output_file, index=False, engine='openpyxl')

# print(f"Translation complete. Translated file saved as {output_file}")


import pandas as pd
import time  # For adding delays
from googletrans import Translator
from requests.exceptions import Timeout, RequestException
from openpyxl import Workbook, load_workbook

# Initialize the translator
translator = Translator()

# Load the Excel file
file_path = 'test.xlsx'  # Change this to your input file path
df = pd.read_excel(file_path, engine='openpyxl')

# Initialize the output Excel file (Create an empty file with headers)
output_file = 'test_output.xlsx'
wb = Workbook()  # Create a new workbook
ws = wb.active

# Write the headers to the output file
ws.append(list(df.columns))

# Save the output file to start writing rows in real-time
wb.save(output_file)

# Function to check if text contains Malayalam characters
def contains_malayalam(text):
    if isinstance(text, str):
        return any('\u0D00' <= char <= '\u0D7F' for char in text)
    return False

# Function to translate text if it contains Malayalam with retry logic
def translate_text(text, retries=3, delay=5):
    
    if contains_malayalam(text):
        for attempt in range(retries):
            try:
                translated = translator.translate(text, src='ml', dest='en')
                return translated.text
            except (Timeout, RequestException) as e:
                print(f"Error translating: {text}. Error: {e}. Attempt {attempt + 1}/{retries}")
                time.sleep(delay)  # Wait before retrying
            except Exception as e:
                print(f"Unexpected error for text: {text}. Error: {e}")
                break
        return text  # Return original if translation fails after retries
    return text  # Return original if not in Malayalam

# Process each row individually and write in real-time
for idx, row in df.iterrows():
    translated_row = []
    
    # Iterate over each cell in the row
    for col in df.columns:
        cell_value = row[col]
        translated_value = translate_text(cell_value)
        translated_row.append(translated_value)
    
    # Load the output file again (to avoid any memory caching)
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Write the translated row to the output file in real-time
    ws.append(translated_row)
    
    # Save the workbook after every row
    wb.save(output_file)
    
    # Display a success message for each row processed
    print(f"Row {idx + 1} processed and written to the output file.")

print(f"Translation complete. Translated file saved as {output_file}")
